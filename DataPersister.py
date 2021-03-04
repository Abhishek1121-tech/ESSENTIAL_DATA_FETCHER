import xlsxwriter
from ConfigReader import ConfigReader
from Constants import Constants
from Utils import Utils
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill

class DataPersister:

    vendor_name_list=['','','','Gournet Garden','Big basket','Big basket discounted','Big basket Organic','Big basket - organic Discounted','Nature\'s Basket','Townness','Townness - Discounted']

    def recDataPersist(data_dict):
        #print(data_dict)
        config_dict=ConfigReader.get_confic_dict()
        status,fileName=DataPersister.createExcelFile()
        if status == Constants.STR_SUCCESS:
            DataPersister.loadExcelAddSheet(fileName,config_dict[Constants.STRING_APP]['daily_excel_sheet_one_name'].replace(Constants.UNDERSCORE,Constants.SPACE))
            DataPersister.firt2RowPersistInSheet(fileName,config_dict[Constants.STRING_APP]['daily_excel_sheet_one_name'].replace(Constants.UNDERSCORE,Constants.SPACE))

    def firt2RowPersistInSheet(file_name,sheet_name):
        workbook=openpyxl.load_workbook(file_name)
        indx_no=DataPersister.makeSheetActive(workbook,sheet_name)
        print(indx_no)
        workbook.active = indx_no
        DataPersister.writeForRow(file_name,workbook)
        #writeForRow1(workbook)
    
    def writeForRow(file_name,workbook):
        row_count = []
        row_fist_value=[Constants.ROW_VALUE_LAST_UPDATE,datetime.today()]
        row_count.append(row_fist_value)
        row_count.append(DataPersister.vendor_name_list)
        print(row_count)
        for count in row_count:
            workbook.active.append(count)
        
        for col_range in range(Constants.NUM_4, len(DataPersister.vendor_name_list)+Constants.NUM_1):
            cell_title = workbook.active.cell(Constants.NUM_2, col_range)
            cell_title.fill = PatternFill(start_color=Constants.FILL_START_COLOR, end_color=Constants.FILL_END_COLOR, fill_type=Constants.FILL_TYPE)
        
        workbook.save(file_name)


    def makeSheetActive(workbook,sheet_name):
        for sheet_indx in range(len(workbook.sheetnames)):
            if workbook.sheetnames[sheet_indx] == sheet_name :
                break
        return sheet_indx
    
    def loadExcelAddSheet(file_name,sheet_name):
        #print(sheet_name)
        #print(fileName)
        workbook=openpyxl.load_workbook(file_name)
        workbook.create_sheet(sheet_name)
        workbook.save(file_name)

    def createExcelFile():
        fileName=DataPersister.AppendFileName()
        #print(fileName)
        if fileName:
            workbook = xlsxwriter.Workbook(fileName)
            workbook.close()
            return Constants.STR_SUCCESS,fileName
        else:
            print("file Name is empty")
            return Constants.EXCEPTION_FILENAME,fileName

    def AppendFileName():
        config_dict=ConfigReader.get_confic_dict()
        excelNameAppender=config_dict[Constants.STRING_APP]['daily_excel_base_path']+config_dict[Constants.STRING_APP]['daily_excel_file_name_prefix'].replace(Constants.UNDERSCORE,Constants.SPACE)+Constants.UNDERSCORE

        if config_dict[Constants.STRING_APP]['daily_excel_appender_type'] == Constants.STRING_DATE:
           excelNameAppender=excelNameAppender+datetime.today().strftime(config_dict[Constants.STRING_APP]['daily_excel_type_format'])
        else:
            print("If appender for the file name changed will add implemetation here in future")

        excelNameAppender=excelNameAppender+Constants.DOT+config_dict[Constants.STRING_APP]['daily_excel_extension']
        return excelNameAppender
import xlsxwriter
from ConfigReader import ConfigReader
from Constants import Constants
from Utils import Utils
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill

class ModelMonthlyBucketDataPersister:

    vendor_name_list=['','SKU(g)','#SKUs','Gournet Garden','Big basket','Big basket discounted','Big basket Organic','Big basket - organic Discounted','Nature\'s Basket','Townness','Townness - Discounted','Namdhari']
    items_persist_in_excel=[Constants.STR_VEG_ORG,Constants.STR_VEG_NATURO,Constants.STR_FRUITS]

    def recDataPersist(data_dict):
        #print(data_dict)
        config_dict=ConfigReader.get_confic_dict()
        status,fileName=ModelMonthlyBucketDataPersister.createExcelFile()
        if status == Constants.STR_SUCCESS:
            ModelMonthlyBucketDataPersister.loadExcelAddSheet(fileName,config_dict[Constants.STRING_APP]['daily_excel_sheet_one_name'].replace(Constants.UNDERSCORE,Constants.SPACE))
            ModelMonthlyBucketDataPersister.firt2RowPersistInSheet(fileName,config_dict[Constants.STRING_APP]['daily_excel_sheet_one_name'].replace(Constants.UNDERSCORE,Constants.SPACE))
            ModelMonthlyBucketDataPersister.writeDictDataInExcelModelMontlyBasket(data_dict,fileName,config_dict[Constants.STRING_APP]['daily_excel_sheet_one_name'].replace(Constants.UNDERSCORE,Constants.SPACE))

    def writeDictDataInExcelModelMontlyBasket(data_dict,file_name,sheet_name):
        #print(data_dict)
        row_counter=3
        for item in ModelMonthlyBucketDataPersister.items_persist_in_excel:
            #print(data_dict[item])
            row_counter=ModelMonthlyBucketDataPersister.persist_item_data_in_excel(item,row_counter,data_dict[item],file_name,sheet_name)
            row_counter+=1

    def persist_item_data_in_excel(item,row_counter,data,file_name,sheet_name):
        cell_counter=Constants.NUM_4
        #print(data)
        workbook=openpyxl.load_workbook(file_name)
        indx_no=ModelMonthlyBucketDataPersister.makeSheetActive(workbook,sheet_name)
        #print(indx_no)
        workbook.active = indx_no
        sheet = workbook.active
        sheet.cell(row=row_counter, column=Constants.NUM_1).value = item
        row_counter+=1
        for d in data:
            sheet.cell(row=row_counter, column=Constants.NUM_1).value = d[Constants.STR_PRODUCT_NAME]
            sheet.cell(row=row_counter, column=Constants.NUM_2).value = d[Constants.STR_SKU_G]
            sheet.cell(row=row_counter, column=Constants.NUM_3).value = d[Constants.STR_SKU_U]
            #print(d)
            for v in ModelMonthlyBucketDataPersister.vendor_name_list[Constants.NUM_3:len(ModelMonthlyBucketDataPersister.vendor_name_list)]:
                #print(v)
                price_counter=0
                for price in d[Constants.STR_VENDOR_LIST]:
                    #print(price)
                    price_counter+=1
                    if price[Constants.NUM_0]==v:
                        sheet.cell(row=row_counter, column=cell_counter).value = price[Constants.NUM_1]
                        cell_counter+=1
                        break
                    if price_counter == len(d[Constants.STR_VENDOR_LIST]):
                        sheet.cell(row=row_counter, column=cell_counter).value = Constants.STR_NOT_AVAILABLE
                        cell_counter+=1
            cell_counter=Constants.NUM_4
            row_counter+=1
        workbook.save(file_name)
        return row_counter



    def firt2RowPersistInSheet(file_name,sheet_name):
        workbook=openpyxl.load_workbook(file_name)
        indx_no=ModelMonthlyBucketDataPersister.makeSheetActive(workbook,sheet_name)
        #print(indx_no)
        workbook.active = indx_no
        ModelMonthlyBucketDataPersister.writeInitialData(file_name,workbook)

    
    def writeInitialData(file_name,workbook):
        row_count = []
        row_fist_value=[Constants.ROW_VALUE_LAST_UPDATE,datetime.today()]
        row_count.append(row_fist_value)
        row_count.append(ModelMonthlyBucketDataPersister.vendor_name_list)
        #print(row_count)
        for count in row_count:
            workbook.active.append(count)
        
        for col_range in range(Constants.NUM_1, len(ModelMonthlyBucketDataPersister.vendor_name_list)+Constants.NUM_1):
            cell_title = workbook.active.cell(Constants.NUM_2, col_range)
            cell_title.fill = PatternFill(start_color=Constants.FILL_START_COLOR, end_color=Constants.FILL_END_COLOR, fill_type=Constants.FILL_TYPE)
        
        workbook.save(file_name)


    def makeSheetActive(workbook,sheet_name):
        for sheet_indx in range(len(workbook.sheetnames)):
            if workbook.sheetnames[sheet_indx] == sheet_name :
                break
        return sheet_indx
    
    def loadExcelAddSheet(file_name,sheet_name):
        #print(sheet_name)
        #print(fileName)
        workbook=openpyxl.load_workbook(file_name)
        workbook.create_sheet(sheet_name)
        workbook.save(file_name)

    def createExcelFile():
        fileName=ModelMonthlyBucketDataPersister.AppendFileName()
        #print(fileName)
        if fileName:
            workbook = xlsxwriter.Workbook(fileName)
            workbook.close()
            return Constants.STR_SUCCESS,fileName
        else:
            print("file Name is empty")
            return Constants.EXCEPTION_FILENAME,fileName

    def AppendFileName():
        config_dict=ConfigReader.get_confic_dict()
        excelNameAppender=config_dict[Constants.STRING_APP]['daily_excel_base_path']+config_dict[Constants.STRING_APP]['daily_excel_file_name_prefix'].replace(Constants.UNDERSCORE,Constants.SPACE)+Constants.UNDERSCORE

        if config_dict[Constants.STRING_APP]['daily_excel_appender_type'] == Constants.STRING_DATE:
           excelNameAppender=excelNameAppender+datetime.today().strftime(config_dict[Constants.STRING_APP]['daily_excel_type_format'])
        else:
            print("If appender for the file name changed will add implemetation here in future")

        excelNameAppender=excelNameAppender+Constants.DOT+config_dict[Constants.STRING_APP]['daily_excel_extension']
        return excelNameAppender